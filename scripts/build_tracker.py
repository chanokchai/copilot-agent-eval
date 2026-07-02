"""Builds the 4-sheet Copilot agent evaluation tracker workbook.

Usage:   python build_tracker.py [output.xlsx]     (default: ./copilot-agent-evaluation-tracker.xlsx)
Requires: pip install openpyxl

To adapt for your own agents, edit three data blocks below:
  - `rows`  (Sheet 1): one entry per agent — name, type, definition, owner, phase data
  - `log`   (Sheet 2): one entry per fix round
  - `items` (Sheet 4): one entry per test question/summary with Pass/Partial/Fail scores
Pass rates on Sheets 1 and 2 are formulas linked to Sheet 4, so they update automatically.
The script prints a reconciliation self-check (pass counts per run) after saving.
"""
import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

ARIAL = "Arial"
DARK = "1F2D3D"
BLUE = "3B6FC4"
GREEN_LINK = "008000"
P0, P1, P2, P3, P4 = "D9E2F3", "DDEBF7", "FCE4D6", "FFF2CC", "E2EFDA"
GRAY = "EEF2F7"
thin = Side(style="thin", color="B9C6D6")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
LOG = "'4. Test Item Log'"

def style(ws, cell, val, bold=False, size=10, fill=None, color="000000", wrap=True, center=False, border=True, fmt=None):
    c = ws[cell]
    c.value = val
    c.font = Font(name=ARIAL, bold=bold, size=size, color=color)
    if fill: c.fill = PatternFill("solid", start_color=fill)
    c.alignment = Alignment(horizontal="center" if center else "left", vertical="center", wrap_text=wrap)
    if border: c.border = BORDER
    if fmt: c.number_format = fmt
    return c

wb = Workbook()

# ---------------- Sheet 1: Agent Tracker ----------------
ws = wb.active
ws.title = "1. Agent Tracker"
ws.sheet_view.showGridLines = False

style(ws, "A1", "Copilot Agent Evaluation — Pilot Tracker", bold=True, size=16, color=DARK, border=False)
style(ws, "A2", "How to use: one row per agent. Fill columns left to right as you finish each phase. Set each phase Status to Done when complete — the Progress column updates automatically. Green text = value calculated from Sheet 4 (Test Item Log); do not type over it.", size=9, color="5A6B7F", border=False)
ws.merge_cells("A2:W2")

groups = [("A3","F3","AGENT INFO",GRAY), ("G3","J3","PHASE 0 — Golden Dataset",P0),
          ("K3","M3","PHASE 1 — Baseline Run",P1), ("N3","O3","PHASE 2 — Failure Mapping",P2),
          ("P3","R3","PHASE 3 — Fix & Re-run",P3), ("S3","U3","PHASE 4 — Pilot & Monitor",P4),
          ("V3","W3","SUMMARY",GRAY)]
for a, b, t, f in groups:
    ws.merge_cells(f"{a}:{b}")
    style(ws, a, t, bold=True, size=10, fill=f, center=True)
    c0 = ws[a].column; c1 = ws[b].column
    for col in range(c0, c1+1):
        ws.cell(row=3, column=col).border = BORDER
        ws.cell(row=3, column=col).fill = PatternFill("solid", start_color=f)

headers = ["No.","Agent Name","Agent Type","Definition (what it does)","Agent Owner","Knowledge Source",
           "# Source Docs","# Test Items","Status","Done Date",
           "Run Date","Baseline Pass %","Status",
           "Top Failure Types Found","Status",
           "# Fix Rounds","Latest Pass %","Status",
           "Pilot Start","User Thumbs-Up %","Status",
           "Progress","Notes / Next Action"]
fills = [GRAY]*6 + [P0]*4 + [P1]*3 + [P2]*2 + [P3]*3 + [P4]*3 + [GRAY]*2
for i, (h, f) in enumerate(zip(headers, fills), 1):
    style(ws, f"{get_column_letter(i)}4", h, bold=True, size=9, fill=f, center=True)

rows = [
 [1,"HR Policy Q&A Bot","Q&A","Answers employee questions about HR policies using the policy handbook","Somchai P.","HR Policy Handbook (SharePoint)",
  10,20,"Done",dt.date(2026,6,5), dt.date(2026,6,10),None,"Done",
  "Made-up facts; does not say \"I don't know\"; wrong chapter retrieved","Done",
  3,None,"Done", dt.date(2026,6,25),0.85,"In Progress",
  None,"Pilot running with 20 users. Review feedback on 10 Jul."],
 [2,"Engineering Spec Summarizer","Summary","Summarizes technical spec documents into the standard 1-page template","Nattaya K.","Engineering specs (Teams folder)",
  12,12,"Done",dt.date(2026,6,12), dt.date(2026,6,16),None,"Done",
  "Wrong output format; numbers copied incorrectly","Done",
  2,None,"In Progress", None,None,"Not Started",
  None,"Round 3: add mandatory-field checklist to the prompt (targets ES-10, ES-12)."],
 [3,"IT Helpdesk Q&A Bot","Q&A","Answers how-to questions about internal IT systems from KB articles","Krit T.","IT Knowledge Base articles",
  10,45,"In Progress",None, None,None,"Not Started",
  None,"Not Started",
  None,None,"Not Started", None,None,"Not Started",
  None,"Writing test questions (2 drafted in Sheet 4, 45 planned incl. trap questions)."],
]
DATE_COLS = {10, 11, 19}; PCT_COLS = {12, 17, 20}
for r, row in enumerate(rows, 5):
    for c, v in enumerate(row, 1):
        cell = style(ws, f"{get_column_letter(c)}{r}", v, size=9,
                     center=c in (1,3,7,8,9,10,11,12,13,15,16,17,18,19,20,21,22))
        if c in DATE_COLS: cell.number_format = "DD-MMM-YY"
        if c in PCT_COLS: cell.number_format = "0%"
    ws[f"V{r}"] = f'=COUNTIF(I{r},"Done")+COUNTIF(M{r},"Done")+COUNTIF(O{r},"Done")+COUNTIF(R{r},"Done")+COUNTIF(U{r},"Done")'
    ws[f"V{r}"].number_format = '0" / 5"'

# pass-% cells pull from Sheet 4 summary (green = intra-workbook link)
links = {"L5": f"={LOG}!C5", "Q5": f"={LOG}!F5", "L6": f"={LOG}!C6", "Q6": f"={LOG}!E6"}
for cell, f in links.items():
    ws[cell] = f
    ws[cell].font = Font(name=ARIAL, size=9, color=GREEN_LINK)
    ws[cell].number_format = "0%"

for r in range(8, 16):
    for c in range(1, 24):
        style(ws, f"{get_column_letter(c)}{r}", None, size=9)
    ws[f"V{r}"] = f'=IF(B{r}="","",COUNTIF(I{r},"Done")+COUNTIF(M{r},"Done")+COUNTIF(O{r},"Done")+COUNTIF(R{r},"Done")+COUNTIF(U{r},"Done"))'
    ws[f"V{r}"].number_format = '0" / 5"'
    for col in (10, 11, 19): ws.cell(row=r, column=col).number_format = "DD-MMM-YY"
    for col in (12, 17, 20): ws.cell(row=r, column=col).number_format = "0%"

dv_status = DataValidation(type="list", formula1='"Not Started,In Progress,Done"', allow_blank=True)
dv_type = DataValidation(type="list", formula1='"Q&A,Summary,Hybrid"', allow_blank=True)
ws.add_data_validation(dv_status); ws.add_data_validation(dv_type)
for col in ("I","M","O","R","U"): dv_status.add(f"{col}5:{col}15")
dv_type.add("C5:C15")

green = PatternFill("solid", start_color="C6EFCE"); yellow = PatternFill("solid", start_color="FFEB9C")
grayf = PatternFill("solid", start_color="E7E6E6"); red = PatternFill("solid", start_color="FFC7CE")
for col in ("I","M","O","R","U"):
    rng = f"{col}5:{col}15"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Done"'], fill=green))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"In Progress"'], fill=yellow))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Not Started"'], fill=grayf))

widths = {"A":5,"B":24,"C":10,"D":40,"E":14,"F":26,"G":9,"H":9,"I":12,"J":11,"K":11,"L":10,"M":12,"N":32,"O":12,"P":8,"Q":10,"R":12,"S":11,"T":11,"U":12,"V":8,"W":42}
for k, v in widths.items(): ws.column_dimensions[k].width = v
ws.row_dimensions[1].height = 24; ws.row_dimensions[2].height = 30; ws.row_dimensions[4].height = 30
for r in range(5, 8): ws.row_dimensions[r].height = 36
ws.freeze_panes = "C5"

# ---------------- Sheet 2: Iteration Log ----------------
ws2 = wb.create_sheet("2. Iteration Log")
ws2.sheet_view.showGridLines = False
style(ws2, "A1", "Iteration Log — one row per fix round (Phase 3)", bold=True, size=14, color=DARK, border=False)
style(ws2, "A2", "Rule: change only ONE thing per round, re-run all golden-dataset tests, and record the result here. Green text = pass rates calculated from Sheet 4 (Test Item Log); do not type over them.", size=9, color="5A6B7F", border=False)
ws2.merge_cells("A2:I2")

h2 = ["Date","Agent Name","Round #","The ONE Change Made","Failure It Targets","Pass % Before","Pass % After","Change (+/-)","Keep Change?"]
for i, h in enumerate(h2, 1): style(ws2, f"{get_column_letter(i)}3", h, bold=True, size=9, fill=P3, center=True)

log = [
 [dt.date(2026,6,11),"HR Policy Q&A Bot",1,'Added rule: "Answer ONLY from the provided sources"',"Made-up facts", "C5","D5"],
 [dt.date(2026,6,13),"HR Policy Q&A Bot",2,'Added instruction + example of saying "I don\'t know"',"No refusal on out-of-scope (trap) questions", "D5","E5"],
 [dt.date(2026,6,17),"HR Policy Q&A Bot",3,"Split the 200-page handbook into chapter files","Wrong document retrieved", "E5","F5"],
 [dt.date(2026,6,18),"Engineering Spec Summarizer",1,"Added one full example of the ideal summary to the prompt","Wrong output format", "C6","D6"],
 [dt.date(2026,6,22),"Engineering Spec Summarizer",2,'Added rule: "Copy all numbers, units and codes exactly"',"Numbers copied incorrectly", "D6","E6"],
]
for r, row in enumerate(log, 4):
    for c, v in enumerate(row[:5], 1):
        cell = style(ws2, f"{get_column_letter(c)}{r}", v, size=9, center=c in (1,3))
        if c == 1: cell.number_format = "DD-MMM-YY"
    for c, ref in ((6, row[5]), (7, row[6])):
        cell = style(ws2, f"{get_column_letter(c)}{r}", f"={LOG}!{ref}", size=9, center=True, color=GREEN_LINK)
        cell.number_format = "0%"
    style(ws2, f"H{r}", f"=G{r}-F{r}", size=9, center=True, fmt="+0%;-0%")
    style(ws2, f"I{r}", f'=IF(H{r}>0,"Yes","Review")', size=9, center=True)
for r in range(9, 16):
    for c in range(1, 10): style(ws2, f"{get_column_letter(c)}{r}", None, size=9)
    ws2[f"H{r}"] = f'=IF(G{r}="","",G{r}-F{r})'; ws2[f"H{r}"].number_format = "+0%;-0%"
    ws2[f"A{r}"].number_format = "DD-MMM-YY"
    ws2[f"F{r}"].number_format = "0%"; ws2[f"G{r}"].number_format = "0%"

dv_keep = DataValidation(type="list", formula1='"Yes,No,Review"', allow_blank=True)
ws2.add_data_validation(dv_keep); dv_keep.add("I9:I15")
for k, v in {"A":11,"B":24,"C":8,"D":48,"E":34,"F":11,"G":11,"H":11,"I":12}.items(): ws2.column_dimensions[k].width = v
ws2.row_dimensions[2].height = 28; ws2.row_dimensions[3].height = 28
ws2.freeze_panes = "A4"

# ---------------- Sheet 3: Scoring Checklist ----------------
ws3 = wb.create_sheet("3. Scoring Checklist")
ws3.sheet_view.showGridLines = False
style(ws3, "A1", "Scoring Checklist — use these questions to score every test item in Sheet 4", bold=True, size=14, color=DARK, border=False)

style(ws3, "A3", "How to score (3-point scale)", bold=True, size=11, fill=GRAY)
ws3.merge_cells("A3:C3")
guide = [("Pass","Meets the criteria fully — nothing to fix."),
         ("Partial","Mostly OK but something needs editing before use."),
         ("Fail","Wrong, missing, or made-up — cannot be used. Any made-up fact = automatic Fail.")]
for r, (k, v) in enumerate(guide, 4):
    style(ws3, f"A{r}", k, bold=True, size=10, center=True,
          fill={"Pass":"C6EFCE","Partial":"FFEB9C","Fail":"FFC7CE"}[k])
    style(ws3, f"B{r}", v, size=10); ws3.merge_cells(f"B{r}:C{r}")
    ws3[f"C{r}"].border = BORDER

def checklist(start, title, fill, items, example):
    style(ws3, f"A{start}", title, bold=True, size=12, fill=fill, color="FFFFFF")
    ws3.merge_cells(f"A{start}:D{start}")
    for col in "BCD": ws3[f"{col}{start}"].fill = PatternFill("solid", start_color=fill); ws3[f"{col}{start}"].border = BORDER
    hdr = start + 1
    for c, h in enumerate(["#","Question to ask about the output","What a FAIL looks like","Score (example)"], 1):
        style(ws3, f"{get_column_letter(c)}{hdr}", h, bold=True, size=9, fill=GRAY, center=True)
    for i, (q, bad, sc) in enumerate(items, 1):
        r = hdr + i
        style(ws3, f"A{r}", i, size=9, center=True)
        style(ws3, f"B{r}", q, size=9)
        style(ws3, f"C{r}", bad, size=9)
        style(ws3, f"D{r}", sc, size=9, center=True,
              fill={"Pass":"C6EFCE","Partial":"FFEB9C","Fail":"FFC7CE"}.get(sc))
        ws3.row_dimensions[r].height = 30
    style(ws3, f"A{hdr+len(items)+1}", example, size=9, color="5A6B7F", border=False)
    ws3.merge_cells(f"A{hdr+len(items)+1}:D{hdr+len(items)+1}")
    return hdr + len(items) + 3

qa = [("Did it find the right document / section?","Cites or clearly uses the wrong source","Pass"),
      ("Is the answer factually correct vs. the source?","Contradicts what the document says","Pass"),
      ("Is every claim backed by the source? (no made-up facts)","Adds details that are not in any document","Fail"),
      ("Does it say \"I don't know\" when the answer is not in our documents?","Invents an answer to an out-of-scope question","Pass"),
      ("Are all parts of a multi-part question answered?","Answers only the first half of the question","Partial")]
nxt = checklist(8, "Q&A AGENTS — score every test answer against these 5 checks", BLUE, qa,
      "Example above: this test item scored Pass, Pass, Fail, Pass, Partial → record it in Sheet 4 as a Fail overall (made-up fact = automatic Fail) and tag the failure type \"Made-up facts\". This is exactly how item HR-14 was scored at baseline.")

summ = [("Does the output follow the required format / template?","Sections missing or in the wrong order","Pass"),
        ("Are all key facts from the source included?","A mandatory spec item is missing","Partial"),
        ("Are numbers, units and codes copied exactly?","A value, unit or part code is wrong (zero tolerance)","Pass"),
        ("Does the engineering logic (cause → effect) still make sense?","Constraints or reasoning are garbled by the summary","Pass"),
        ("Would you send it to a colleague without editing?","You would need to rewrite parts before sharing","Pass")]
checklist(nxt, "SUMMARY AGENTS — score every test summary against these 5 checks", "5B4FB5", summ,
      "Example above: 4 Pass + 1 Partial → record in Sheet 4 as Partial overall and tag the failure type \"Missing details\". This is exactly how item ES-10 was scored in Round 1.")

dv_score = DataValidation(type="list", formula1='"Pass,Partial,Fail"', allow_blank=True)
ws3.add_data_validation(dv_score)
dv_score.add("D10:D14"); dv_score.add(f"D{nxt+2}:D{nxt+6}")
for k, v in {"A":9,"B":52,"C":44,"D":16}.items(): ws3.column_dimensions[k].width = v

# ---------------- Sheet 4: Test Item Log ----------------
ws4 = wb.create_sheet("4. Test Item Log")
ws4.sheet_view.showGridLines = False
style(ws4, "A1", "Test Item Log — one row per test question / summary, scored on every run", bold=True, size=14, color=DARK, border=False)
style(ws4, "A2", "How to use: write test items during Phase 0. Score every item (Pass / Partial / Fail — use the checklist in Sheet 3) at baseline and again after each fix round. The pass-rate table below updates automatically and feeds Sheets 1 and 2.", size=9, color="5A6B7F", border=False)
ws4.merge_cells("A2:L2")

# summary table (feeds Sheets 1 & 2)
style(ws4, "A4", "Agent", bold=True, size=9, fill=GRAY, center=True)
style(ws4, "B4", "# Test Items", bold=True, size=9, fill=GRAY, center=True)
style(ws4, "C4", "Baseline Pass %", bold=True, size=9, fill=P1, center=True)
style(ws4, "D4", "Round 1 Pass %", bold=True, size=9, fill=P3, center=True)
style(ws4, "E4", "Round 2 Pass %", bold=True, size=9, fill=P3, center=True)
style(ws4, "F4", "Round 3 Pass %", bold=True, size=9, fill=P3, center=True)
style(ws4, "G4", "Pass % = items scored Pass ÷ items scored (Partial and Fail both count as not passed). Feeds Sheets 1 and 2 automatically.", size=8, color="5A6B7F", border=False)
ws4.merge_cells("G4:L4")

agents = ["HR Policy Q&A Bot","Engineering Spec Summarizer","IT Helpdesk Q&A Bot"]
DATA_TOP, DATA_BOT = 13, 54
for r, name in enumerate(agents, 5):
    style(ws4, f"A{r}", name, size=9)
    style(ws4, f"B{r}", f'=COUNTIF($B${DATA_TOP}:$B${DATA_BOT},$A{r})', size=9, center=True)
    for col, sc in zip("CDEF", "GHIJ"):
        f = (f'=IF(COUNTIFS($B${DATA_TOP}:$B${DATA_BOT},$A{r},{sc}${DATA_TOP}:{sc}${DATA_BOT},"<>")=0,"",'
             f'COUNTIFS($B${DATA_TOP}:$B${DATA_BOT},$A{r},{sc}${DATA_TOP}:{sc}${DATA_BOT},"Pass")/'
             f'COUNTIFS($B${DATA_TOP}:$B${DATA_BOT},$A{r},{sc}${DATA_TOP}:{sc}${DATA_BOT},"<>"))')
        style(ws4, f"{col}{r}", f, size=9, center=True, bold=True, fmt="0%")

style(ws4, "A9", "TEST ITEMS", bold=True, size=11, fill=DARK, color="FFFFFF")
ws4.merge_cells("A9:L9")
for col in "BCDEFGHIJKL":
    ws4[f"{col}9"].fill = PatternFill("solid", start_color=DARK); ws4[f"{col}9"].border = BORDER
style(ws4, "A10", "Score meaning: Pass = usable as-is · Partial = needs edits · Fail = wrong / made-up (see Sheet 3). Leave a score blank if the run has not happened yet.", size=8.5, color="5A6B7F", border=False)
ws4.merge_cells("A10:L10")

h4 = ["Item ID","Agent Name","Source Document","Question / Task","Expected Answer / Reference","Trap?",
      "Baseline","Round 1","Round 2","Round 3","Latest Failure Tag","Notes"]
f4 = [GRAY]*6 + [P1] + [P3]*3 + [P2, GRAY]
HR_ROW = 12
for i, (h, f) in enumerate(zip(h4, f4), 1):
    style(ws4, f"{get_column_letter(i)}{HR_ROW}", h, bold=True, size=9, fill=f, center=True)

P, PA, F = "Pass", "Partial", "Fail"
hr = "HR Policy Q&A Bot"; es = "Engineering Spec Summarizer"; it = "IT Helpdesk Q&A Bot"
items = [
 # --- HR Policy Q&A Bot: 20 items. Baseline 13/20=65%, R1 16/20=80%, R2 18/20=90%, R3 19/20=95%
 ("HR-01",hr,"Ch.1 Leave Policy","How many days of annual leave do new employees get?","10 days in the first year","No",P,P,P,P,"",""),
 ("HR-02",hr,"Ch.1 Leave Policy","Can unused annual leave be carried over to next year?","Yes, maximum 5 days","No",P,P,P,P,"",""),
 ("HR-03",hr,"Ch.2 Benefits","What is the dental allowance per year?","5,000 THB per year","No",P,P,P,P,"",""),
 ("HR-04",hr,"Ch.2 Benefits","Who is eligible for group health insurance?","All full-time employees from start date","No",P,P,P,P,"",""),
 ("HR-05",hr,"Ch.3 Payroll","When is salary paid each month?","On the 25th of each month","No",P,P,P,P,"",""),
 ("HR-06",hr,"Ch.3 Payroll","How do I request a salary certificate?","Via HR portal; issued within 3 working days","No",P,P,P,P,"",""),
 ("HR-07",hr,"Ch.4 Code of Conduct","What is the policy on accepting gifts from vendors?","Declare all gifts; not allowed above 3,000 THB","No",P,P,P,P,"",""),
 ("HR-08",hr,"Ch.5 Travel & Expense","What is the daily meal allowance for domestic trips?","800 THB per day","No",P,P,P,P,"",""),
 ("HR-09",hr,"Ch.5 Travel & Expense","Which hotels can be booked for business trips?","Approved list, max 2,500 THB per night","No",P,P,P,P,"",""),
 ("HR-10",hr,"Ch.6 Probation","How long is the probation period?","119 days","No",P,P,P,P,"",""),
 ("HR-11",hr,"Ch.7 Overtime","What is the OT rate on public holidays?","3x hourly rate","No",P,P,P,P,"",""),
 ("HR-12",hr,"Ch.8 Remote Work","How many remote days per week are allowed?","Up to 2 days with manager approval","No",P,P,P,P,"",""),
 ("HR-13",hr,"Ch.9 Training","What is the annual training budget per employee?","15,000 THB per year","No",P,P,P,P,"",""),
 ("HR-14",hr,"Ch.2 Benefits","Does the company match provident fund contributions?","Yes, 5% of base salary","No",F,P,P,P,"","Baseline invented a 10% rate — fixed by Round 1 grounding rule"),
 ("HR-15",hr,"Ch.10 Termination","What is the notice period for resignation?","30 days written notice","No",F,P,P,P,"","Baseline said 15 days (made up) — fixed by Round 1"),
 ("HR-16",hr,"Ch.1 Leave Policy","How do I request sick leave and what documents are needed?","Notify manager on day 1; medical certificate if more than 3 days","No",F,P,P,P,"","Baseline answered only half + added invented steps — fixed by Round 1"),
 ("HR-17",hr,"(not in handbook)","What is the company car policy for managers?","Not covered — agent must say \"I don't know\"","Yes",F,F,P,P,"","Trap question. Invented an answer until Round 2 refusal example"),
 ("HR-18",hr,"(not in handbook)","Can I buy extra vacation days?","Not covered — agent must say \"I don't know\"","Yes",F,F,P,P,"","Trap question. Fixed by Round 2"),
 ("HR-19",hr,"Ch.7 Overtime","How is weekend OT calculated, and does it apply to supervisors?","1.5x rate; supervisors are exempt","No",PA,PA,PA,PA,"Missing details","Always answers the rate but skips the supervisor part — candidate for Round 4"),
 ("HR-20",hr,"Ch.8 Remote Work","What equipment does the company provide for remote work?","Laptop + monitor on request","No",F,F,F,P,"","Kept retrieving the Travel chapter — fixed by Round 3 handbook split"),
 # --- Engineering Spec Summarizer: 12 items. Baseline 7/12=58%, R1 9/12=75%, R2 10/12=83%
 ("ES-01",es,"SPEC-001 Pump Assembly","Summarize into standard 1-page template","Approved reference summary ES-01","No",P,P,P,None,"",""),
 ("ES-02",es,"SPEC-002 Motor Control Unit","Summarize into standard 1-page template","Approved reference summary ES-02","No",P,P,P,None,"",""),
 ("ES-03",es,"SPEC-003 Conveyor Drive","Summarize into standard 1-page template","Approved reference summary ES-03","No",P,P,P,None,"",""),
 ("ES-04",es,"SPEC-004 HVAC Layout","Summarize into standard 1-page template","Approved reference summary ES-04","No",P,P,P,None,"",""),
 ("ES-05",es,"SPEC-005 Piping Isometric","Summarize into standard 1-page template","Approved reference summary ES-05","No",P,P,P,None,"",""),
 ("ES-06",es,"SPEC-006 Electrical Panel","Summarize into standard 1-page template","Approved reference summary ES-06","No",P,P,P,None,"",""),
 ("ES-07",es,"SPEC-007 PLC I/O List","Summarize into standard 1-page template","Approved reference summary ES-07","No",P,P,P,None,"",""),
 ("ES-08",es,"SPEC-008 Safety Interlock","Summarize into standard 1-page template","Approved reference summary ES-08","No",F,P,P,None,"","Baseline ignored the template sections — fixed by Round 1 example"),
 ("ES-09",es,"SPEC-009 Hydraulic Press","Summarize into standard 1-page template","Approved reference summary ES-09","No",F,P,P,None,"","Baseline wrong section order — fixed by Round 1"),
 ("ES-10",es,"SPEC-010 Cooling Tower","Summarize into standard 1-page template","Approved reference summary ES-10","No",F,PA,PA,None,"Missing details","Format fixed in Round 1 but maintenance section still too thin — target of Round 3 checklist"),
 ("ES-11",es,"SPEC-011 Air Compressor","Summarize into standard 1-page template","Approved reference summary ES-11","No",F,F,P,None,"","Pressure value copied wrong at baseline & R1 — fixed by Round 2 exact-copy rule"),
 ("ES-12",es,"SPEC-012 Lighting Plan","Summarize into standard 1-page template","Approved reference summary ES-12","No",F,F,F,None,"Numbers copied wrong","Converts lux values incorrectly — target of Round 3"),
 # --- IT Helpdesk Q&A Bot: Phase 0 in progress, no runs yet
 ("IT-01",it,"KB-0012 VPN Setup","How do I connect to the VPN from home?","Install GlobalConnect, login with AD account","No",None,None,None,None,"","Phase 0 draft — not tested yet (45 items planned)"),
 ("IT-02",it,"(not in KB)","Can I install personal software on my work laptop?","Not covered — agent must say \"I don't know\"","Yes",None,None,None,None,"","Phase 0 draft trap question — not tested yet"),
]
for r, row in enumerate(items, DATA_TOP):
    for c, v in enumerate(row, 1):
        style(ws4, f"{get_column_letter(c)}{r}", v, size=9, center=c in (1,6,7,8,9,10))
    ws4.row_dimensions[r].height = 26
for r in range(DATA_TOP + len(items), DATA_BOT + 1):
    for c in range(1, 13): style(ws4, f"{get_column_letter(c)}{r}", None, size=9)

dv_score4 = DataValidation(type="list", formula1='"Pass,Partial,Fail"', allow_blank=True)
dv_trap = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
ws4.add_data_validation(dv_score4); ws4.add_data_validation(dv_trap)
for col in "GHIJ": dv_score4.add(f"{col}{DATA_TOP}:{col}{DATA_BOT}")
dv_trap.add(f"F{DATA_TOP}:F{DATA_BOT}")
for col in "GHIJ":
    rng = f"{col}{DATA_TOP}:{col}{DATA_BOT}"
    ws4.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Pass"'], fill=green))
    ws4.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Partial"'], fill=yellow))
    ws4.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Fail"'], fill=red))

for k, v in {"A":9,"B":24,"C":22,"D":44,"E":40,"F":7,"G":10,"H":10,"I":10,"J":10,"K":16,"L":48}.items():
    ws4.column_dimensions[k].width = v
ws4.row_dimensions[2].height = 30; ws4.row_dimensions[10].height = 14; ws4.row_dimensions[12].height = 26
ws4.freeze_panes = f"C{DATA_TOP}"

import sys
wb.save(sys.argv[1] if len(sys.argv) > 1 else "copilot-agent-evaluation-tracker.xlsx")

# ---- reconciliation self-check (mirrors the COUNTIFS logic) ----
from collections import defaultdict
runs = defaultdict(lambda: defaultdict(lambda: [0, 0]))
for row in items:
    for run_i, col in enumerate(row[6:10]):
        if col is not None:
            runs[row[1]][run_i][1] += 1
            if col == "Pass": runs[row[1]][run_i][0] += 1
for a in runs:
    print(a, [(f"run{k}", f"{v[0]}/{v[1]}", round(v[0]/v[1]*100)) for k, v in sorted(runs[a].items())])
